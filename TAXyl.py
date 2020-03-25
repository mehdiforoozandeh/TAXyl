import warnings
import sys

warnings.filterwarnings("ignore")
warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=DeprecationWarning)

if not sys.warnoptions:
    warnings.simplefilter("ignore")

import openpyxl as xl
import pandas as pd
import numpy as np
import datetime
from pypro import PyPro
from sklearn.model_selection import train_test_split
import joblib
from sklearn.preprocessing import MinMaxScaler
import matplotlib.pyplot as plt

st = datetime.datetime.now()

# ===================== DATA PREP ======================= #

datasetii = xl.load_workbook('DsIV.xlsx')
df = datasetii.active

templabel = pd.DataFrame(index=list(range(1, df.max_row + 1)), columns=['entry', 'templabel', 'seq'])

for i in range(1, df.max_row + 1):
    if 0 <= float(df.cell(row=i, column=2).value) <= 100:
        templabel['entry'][i] = df.cell(row=i, column=1).value
        templabel['templabel'][i] = float(df.cell(row=i, column=2).value)
        templabel['seq'][i] = str(df.cell(row=i, column=3).value)

# Remove unknown amino acids

for i in range(1, len(templabel) + 1):
    if 'X' in templabel['seq'][i]:
        seq = templabel['seq'][i]
        templabel['seq'][i] = seq[:seq.find('X')] + seq[seq.find('X') + 1:]
        # print len(seq), len(templabel['seq'][i])

print(templabel)

# FEATURE EXTRACTION
prot = PyPro()
desc_name = ['AAC', '2AAC', '3AAC', 'CTD', 'PAAC', 'APAAC', 'Moran', 'Geary',
             'MoreauBroto', 'CTF', 'QSO', 'SOCN']
descriptor = [prot.GetAAComp, prot.GetDPComp, prot.GetTPComp, prot.GetCTD,
              prot.GetPAAC, prot.GetAPAAC,
              prot.GetMoranAuto, prot.GetGearyAuto, prot.GetMoreauBrotoAuto,
              prot.GetTriad, prot.GetQSO, prot.GetSOCN]
for name, desc in zip(desc_name, descriptor):
    templabel[name] = pd.Series(np.zeros(len(templabel)), index=templabel.index)
    templabel[name] = templabel[name].astype(object)
    print('Currently Calculating ' + str(name))
    for i in range(1, len(templabel) + 1):
        prot.ReadProteinSequence(templabel['seq'][i])
        if desc == prot.GetPAAC:
            t = desc(lamda=15, weight=0.05)
        elif desc == prot.GetAPAAC:
            t = desc(lamda=15, weight=0.5)
        else:
            t = desc()
        t = list(t.values())
        templabel.set_value(i, name, list(t))

i = 1
dimen = len(np.array(templabel['AAC'][i])) + len(np.array(templabel['2AAC'][i])) + len(
    np.array(templabel['3AAC'][i])) + len(np.array(templabel['CTD'][i])) + len(np.array(templabel['PAAC'][i])) + len(
    np.array(templabel['APAAC'][i])) + len(np.array(templabel['Moran'][i])) + len(
    np.array(templabel['Geary'][i])) + len(np.array(templabel['MoreauBroto'][i])) + len(
    np.array(templabel['CTF'][i])) + len(np.array(templabel['QSO'][i])) + len(np.array(templabel['SOCN'][i]))
# 20 400 8000 147 25 30 240 240 240 512 100 90 (sum=10044)

l = []
for i in range(1, len(templabel) + 1):
    l.append(templabel['AAC'][i] + templabel['2AAC'][i] + templabel['3AAC'][i] + templabel['CTD'][i]
             + templabel['PAAC'][i] + templabel['APAAC'][i] + templabel['Moran'][i] + templabel['Geary'][i] +
             templabel['MoreauBroto'][i]
             + templabel['CTF'][i] + templabel['QSO'][i] + templabel['SOCN'][i])

dataset = np.array(l)
dataset = np.reshape(dataset, [-1, dimen])

# Remove Duplicate Columns

_, idx = np.unique(dataset, axis=1, return_index=True)
dataset = dataset[:, np.sort(idx)]
# print 'idx', idx
print('num of duplicate features:   ' + str(dimen - dataset.shape[1]))
idxstr = ''
for i in idx:
    idxstr += str(i) + ','
idxstr = idxstr[:-1]
# print 'idxstr', idxstr
with open('tempdupidx.txt', 'w') as conf:
    conf.write(idxstr)

# Normalization
scaler = MinMaxScaler(feature_range=(0, 1))
scaler.fit(dataset)
dataset = scaler.transform(dataset)

ylabel = []

# for i in range(1, len(templabel) + 1):
#     if int(templabel['templabel'][i]) < 50:
#         ylabel.append('NonThermo')
#     elif 50 <= int(templabel['templabel'][i]):
#         ylabel.append('Thermo')

# key = {'NonThermo': 0, 'Thermo': 1}
# invkey = {v: k for k, v in list(key.items())}

for i in range(1, len(templabel) + 1):
    if int(templabel['templabel'][i]) <= 50:
        ylabel.append('Meso')
    elif 50 < int(templabel['templabel'][i]) < 75:
        ylabel.append('Thermo')
    elif 75 <= int(templabel['templabel'][i]):
        ylabel.append('HyperThermo')

key = {'Meso': 0, 'Thermo': 1, 'HyperThermo': 2}
invkey = {v: k for k, v in key.items()}


for i in range(len(ylabel)):
    if ylabel[i] in list(key.keys()):
        ylabel[i] = key[ylabel[i]]

ylabel = np.reshape(np.array(ylabel), [-1, 1])

from sklearn.ensemble import RandomForestClassifier
RF_model = RandomForestClassifier(n_estimators=1000, criterion='entropy')

# Define feature selectors
from sklearn.feature_selection import SelectFdr
from sklearn.feature_selection import SelectKBest, chi2


# evalmcic = eval(dataset, ylabel)

# evalmcic.to_csv("EVALTAXyl.csv")

# curve_generator()

def ModelSelection(x, y, K):
    from sklearn.model_selection import KFold
    from sklearn.neural_network import MLPClassifier
    from sklearn.neighbors import KNeighborsClassifier
    from sklearn.svm import SVC
    from sklearn.gaussian_process import GaussianProcessClassifier
    from sklearn.gaussian_process.kernels import RBF
    from sklearn.tree import DecisionTreeClassifier
    from sklearn.ensemble import RandomForestClassifier, AdaBoostClassifier
    from sklearn.naive_bayes import GaussianNB
    from sklearn.discriminant_analysis import QuadraticDiscriminantAnalysis
    names = ["RBF SVM", "Gini Random Forest 1000", 
            'Entropy Random Forest 1000', "Neural Net 3L", "AdaBoost", "Naive Bayes"]
    classifiers = [
        SVC(kernel="rbf", C=211, decision_function_shape='ovo'),
        RandomForestClassifier(n_estimators=1000),
        RandomForestClassifier(n_estimators=1000, criterion='entropy'),
        MLPClassifier(hidden_layer_sizes=(100,100,100), alpha=0.6),
        AdaBoostClassifier(),
        GaussianNB()]
    
    kf = KFold(n_splits=6, shuffle=True, random_state=47)
    accdict = {}
    for name, clf in zip(names, classifiers):
        acc = []
        for train_index, test_index in kf.split(x):
            X_train, X_test = x[train_index], x[test_index]
            y_train, y_test = ylabel[train_index], ylabel[test_index]
            
            selector = SelectKBest(k= K)
            selector.fit(X_train, y_train)
            X_train = selector.transform(X_train)
            X_test = selector.transform(X_test)

            clf.fit(X_train, y_train)
            score = clf.score(X_test, y_test)
            acc.append(score)

        accdict[str(name)] = np.mean(acc)

    return accdict


def save():
    X_train, X_test, y_train, y_test = train_test_split(dataset, ylabel, test_size=0.01, shuffle=True)

    selector = SelectKBest(k= 400)
    selector.fit(X_train, y_train)

    X_train = selector.transform(X_train)
    X_test = selector.transform(X_test)


    RF_model.fit(X_train, y_train)
    print('The final models score: ', RF_model.score(X_test, y_test))

    joblib.dump(scaler, 'Scaler.sav')
    joblib.dump(selector, 'Selector.sav')
    joblib.dump(RF_model, 'RF_model.sav')

save()

# def curve_generator():
#     from sklearn.metrics import auc, roc_curve
#     X_train, X_test, y_train, y_test = train_test_split(dataset, ylabel, test_size=0.1, shuffle=True)

#     selector = SelectKBest(k= 400)
#     selector.fit(X_train, y_train)

#     X_train = selector.transform(X_train)
#     X_test = selector.transform(X_test)
#     RF_model.fit(X_train, y_train)

#     fpr, tpr, threshold = roc_curve(y_test, RF_model.predict(X_test))
#     roc_auc = auc(fpr, tpr)

#     plt.title('Receiver Operating Characteristic')
#     plt.plot(fpr, tpr, 'b', label = 'AUC = %0.2f' % roc_auc)
#     plt.legend(loc = 'lower right')
#     plt.plot([0, 1], [0, 1],'r--')
#     plt.xlim([0, 1])
#     plt.ylim([0, 1])
#     plt.ylabel('True Positive Rate')
#     plt.xlabel('False Positive Rate')
#     plt.show()

# import matplotlib
# colors = matplotlib.colors.cnames


# ress = {"RBF SVM": [], 
#         "Gini Random Forest 1000":[],
#         'Entropy Random Forest 1000':[], 
#         "Neural Net 3L": [], "AdaBoost": [], "Naive Bayes": []}

# ks = range(5, int(dataset.shape[1]/3), 40)
# for i in ks:
#     print(i)
#     res = ModelSelection(dataset, ylabel, i)
#     print(res)
#     maxx = max(res.values())
#     print(maxx)
#     for j in res.keys():
#         if res[j] == maxx:
#             print('best clf: ', j)

#     for k in res.keys():
#         ress[k].append(res[k])


# for k, v in ress.items():
#     rand = np.random.randint(len(colors))
#     plt.plot(ks, v, label = str(k))

# plt.legend()
# plt.show()

